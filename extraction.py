import pandas as pd
import os
import requests
import json
import time
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv
load_dotenv()

"""
Define all global variables
"""
API_KEY = os.getenv('pv_api_key')
BASE_URL = 'https://search.patentsview.org'
TIMEOUT_CODE = 429
ERROR_CODE = 200
ASSIGNEE_TYPE_DICT = { # Dictionary for converting assignee type into it's actual meaning
    1: "Unassigned",
    2: "United States company or corporation",
    3: "Foreign company or corporation",
    4: "United States individual",
    5: "Foreign individual",
    6: "U.S. Federal government",
    7: "Foreign government",
    8: "U.S. county government",
    9: "U.S. state government"
}
PATENT_FIELDS = ["patent_id", "patent_title", "patent_abstract", "patent_date", "patent_type"]

"""
Download all CPC subclass titles from API and create a lookup dictionary
"""
def cpc_subclass_dict():
    endpoint = 'api/v1/cpc_subclass'
    data = {'size': 1000}
    responses = []

    # Continue extracting until responses length equals total hits
    while True:
        response = requests.get(f"{BASE_URL}/{endpoint}/?o={json.dumps(data)}", headers={"X-Api-Key": API_KEY})

        # Check for errors
        if response.status_code == 429:
            wait_for = response.headers['Retry-After']
            time.sleep(wait_for)
            continue
        if response.status_code != 200:
            raise Exception(response.headers)
        
        # Save data
        response_data = response.json()
        responses += response_data['cpc_subclasses']
        expected = response_data['total_hits']
        data['after'] = response_data['cpc_subclasses'][-1]['cpc_subclass_id']

        # End case
        if (response_data['count'] == 0) or (len(responses) >= expected):
            break
    return {item['cpc_subclass_id']: item['cpc_subclass_title'] for item in responses}

CPC_SUBCLASS_DICT = cpc_subclass_dict()

"""
Get all extraction output from the PV API

Requires that base_url, endpoint, and api_key are all defined
"""
def full_extraction_output(assignee_IDs, simplified):
    # Determine return fields
    if simplified:
        f_list = ["patent_id", "patent_title", "patent_abstract", "patent_date", "patent_type", "assignees.*"]
    else:
        ["patent_id", "patent_title", "patent_abstract", "patent_date", "patent_type", "assignees.*",\
            "inventors.*", "cpc_current.*"]
    
    full_output = []
    endpoint = 'api/v1/patent'
    param_dict = {
        "f" : f_list,
        "o" : {"size":1000},
        "q" : {"assignees.assignee_id": assignee_IDs},
        "s" : [{"patent_id":"asc"}],
    }
    
    while True:
        # Update query parameters and get response
        param_string = "&".join([f"{param_name}={json.dumps(param_val)}" for param_name, param_val in param_dict.items()])
        query_url = f"{BASE_URL}/{endpoint.strip('/')}/?{param_string}"
        response = requests.get(query_url, headers={"X-Api-Key": API_KEY})

        # Check for errors
        if response.status_code == 429:
            wait_for = response.headers['Retry-After']
            time.sleep(wait_for)
            continue
        if response.status_code != 200:
            raise Exception(response.headers)

        # Save data
        response_data = response.json()
        full_output += response_data['patents']
        expected = response_data['total_hits']
        param_dict["o"]['after'] = response_data['patents'][-1]['patent_id']

        # End case
        if (response_data['count'] == 0) or (len(full_output) >= expected):
            break 
    return full_output

def new_assignees(row, assignee_IDs):
    if 'assignees' in row.keys():
        new_assignees = []
        for assignee in row['assignees']:
            assignee['assignee'] = assignee['assignee'][47:-1]
            if assignee['assignee'] in assignee_IDs:
                assignee['assignee_type'] = ASSIGNEE_TYPE_DICT[int(assignee['assignee_type'])]
                new_assignees.append(assignee)
        return new_assignees
    else: # Empty assignees
        return [{'assignee': '', 'assignee_type': '', 'assignee_individual_name_first': '', 'assignee_individual_name_last': '',\
            'assignee_organization': '', 'assignee_city': '', 'assignee_state': '', 'assignee_country': '', 'assignee_sequence': ''}]

def new_inventors(row):
    # Extract disamb_ID from URL for inventors
    if 'inventors' in row.keys():
        inventors = row['inventors'].copy()
        for inventor in inventors:
            inventor['inventor'] = inventor['inventor'][47:-1]
        return inventors
    else: # Empty inventors
        return [{'inventor': '', 'inventor_name_first': '', 'inventor_name_last': '', 'inventor_city': '',\
            'inventor_state': '', 'inventor_country': '', 'inventor_sequence': ''}]

def new_cpc(row):
    # Handle cases with empty cpc_current   
    if 'cpc_current' in row.keys():
        new_cpc = []
        subclasses_observed = set()
        for cpc in row['cpc_current']:
            subclass_id = cpc['cpc_subclass_id']
            if subclass_id not in subclasses_observed:
                subclasses_observed.add(subclass_id)
                new_cpc.append({'cpc_section': subclass_id[0], 'cpc_subclass_id': subclass_id, 'cpc_subclass_title': CPC_SUBCLASS_DICT[subclass_id]})
        return new_cpc
    else:
        return [{'cpc_section': '', 'cpc_subclass_id': '', 'cpc_subclass_title': ''}]
    
def process_extraction_output(dirty_output, assignee_IDs, simplified):
    clean_output = []

    for row in dirty_output:
        base_row = {field: row[field] for field in PATENT_FIELDS}
        assignees = new_assignees(row, assignee_IDs)
        if not simplified: # Simplified output doesn't include inventors or cpc
            base_row['inventors'] = new_inventors(row)
            base_row['cpc_current'] = new_cpc(row)

        for assignee in assignees:
            None
            # TODO - create new_row from base_row and append to clean_output
    
    return clean_output


def add_empty_rows(df, max_rows):
    num_empty_rows = max_rows - len(df)
    empty_rows = pd.DataFrame([{}]*num_empty_rows, columns=df.columns)
    return pd.concat([df, empty_rows], ignore_index=True)

def partition_endpoints(result):
    # Create separate dataframes
    inventors = pd.DataFrame.from_dict(result['inventors']).sort_values('inventor_sequence')
    cpc = pd.DataFrame.from_dict(result['cpc_current']).sort_values('cpc_subclass_id')
    assignees = pd.DataFrame.from_dict(result['assignees']).sort_values('assignee_sequence')

    # Add empty rows
    row_count = max(len(inventors.index), len(cpc.index), len(assignees.index))
    inventors = add_empty_rows(inventors, row_count)
    cpc = add_empty_rows(cpc, row_count)
    assignees = add_empty_rows(assignees, row_count)
    
    return inventors, cpc, assignees

def extraction_output_to_csv(output, output_path = "data/05 - extraction/output.csv"):
    df_list = []

    for result in output:
        # Separating data to add empty rows where necessary
        inventors, cpc, assignees = partition_endpoints(result)
        patent = {field: result[field] for field in PATENT_FIELDS}
        patent_df = pd.DataFrame.from_dict([patent] * max(len(inventors.index), len(cpc.index), len(assignees.index)))

        # Combining and storing result
        result = pd.concat([patent_df, assignees, inventors, cpc], axis=1)
        df_list.append(result)
        
    pd.concat(df_list).to_csv(output_path)

def extraction_output_to_excel(output, output_path = "data/05 - extraction/output.xlsx"):
    wb = Workbook()
    ws = wb.active

    for result in output:
        # Separating data to add empty rows where necessary
        inventors, cpc, assignees = partition_endpoints(result)
        patent = {field: result[field] for field in PATENT_FIELDS}
        patent_df = pd.DataFrame.from_dict([patent] * max(len(inventors.index), len(cpc.index), len(assignees.index)))

        # Combining and storing result
        result = pd.concat([patent_df, assignees, inventors, cpc], axis=1)
        merge_start = ws.max_row + 1
        merge_end = merge_start + len(result.index) - 1
        for row in dataframe_to_rows(result, index=False, header=(merge_start == 2)):
            ws.append(row)
        for col in range(1, 6):
            ws.merge_cells(start_row=merge_start, end_row=merge_end, start_column=col, end_column=col)

    wb.save(output_path)

def run_extraction(assignee_IDs=["a0ba1f5c-6e5f-4f62-b309-22bd81c8b043"], output_path=None, simplified=True):
    dirty_output = full_extraction_output(assignee_IDs, simplified)
    clean_output = process_extraction_output(dirty_output, assignee_IDs, simplified)
    if output_path[-4:] == ".csv":
        extraction_output_to_csv(clean_output, output_path)
    elif output_path[-5:] == ".xlsx":
        extraction_output_to_excel(clean_output, output_path)

if __name__ == "__main__":
    obj = run_extraction()
    

# """
# Parameters
# ----------
# assignee_disamiguation_IDs : list[str]
#     List of disambiguated assignee IDs (`assignee_id` field values)
# output_path : str
#     Path to CSV file for saving data for one mention ID (naming convention is simply the mention ID)

# Processing
# ----------
# 1. Pull `df` with SQL connection
# 2. Group-by mention_id in `df` then sort by inventor_id, and cpc_subgroup_id within the groups
# 3. Going group by group, add all rows from `df` to `ws`
# 4. After adding each group, merge all cells that share a value within each column (need to be neighboring cells)
#     - Make sure to not merge cells across separate mention_id groups

# Returns
# -------
# openpyxl Workbook with rows for all assignee mentions that correspond to one disambiguated assignee ID in the 
# provided list. 
# The columns should be the same as in the `assignee_data` function.
# Specifically, rows should be of the form `assignee_data(mention_id)` for each mention_id that corresponds to one of 
# the disambiguated assignee ID in the provided list.
# Implementing merged cells to increase readability
# """
# def disambiguated_assignees_data(assignee_disambiguation_IDs: list[str], connection):
#     id_list = '("' + '","'.join(assignee_disambiguation_IDs) + '")'
#     query = f"SELECT * FROM algorithms_assignee_labeling.assignee a WHERE a.disambiguated_assignee_id IN {id_list}"
#     result = connection.execute(text(query)).fetchall()
#     df = pd.DataFrame(result).drop_duplicates()

#     # Create a new Excel workbook and add a worksheet
#     wb = Workbook()
#     ws = wb.active
    
#     # Loop through mention_id groups and sort values
#     for group, data in tqdm(df.groupby(['patent_id', 'assignee_sequence']), "Outer groupby loop"):
#         df_temp = data.sort_values(by=["inventor_sequence", "cpc_subgroup_id"], ascending=True, inplace=False)
    
#         # Determine indexing values and add rows to worksheets
#         index_start = ws.max_row + 1
#         index_end = index_start + len(df_temp.index) - 1
#         for row in dataframe_to_rows(df_temp, index=False, header=(index_start == 2)):
#             ws.append(row)
    
#         merge_cells(ws, index_start, index_end, len(df.columns))

#     return wb