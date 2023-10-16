from sqlalchemy import create_engine, text
from dotenv import dotenv_values
import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm

"""
Dictionary for converting assignee type into it's actual meaning
"""
ASSIGNEE_TYPE_DICT = {
    1: "Unassigned",
    2: "United States company or corporation",
    3: "Foreign company or corporation",
    4: "United States individual",
    5: "Foreign individual",
    6: "U.S. Federal government",
    7: "Foreign government",
    8: "U.S. county government",
    9: "U.S. state government"
}


"""
Establish MySQL connection with environmental variables

Returns
-------
engine : a PyMySQL engine object
"""
def establish_connection():
    config = dotenv_values(".env")
    user = config['user']
    pswd = config['password']
    hostname = config['hostname']
    dbname = config['dbname']
    engine = create_engine(f"mysql+pymysql://{user}:{pswd}@{hostname}/{dbname}?charset=utf8mb4")
    return engine


"""
Parameters
----------
mention_id : string
    Mention ID of an assignee, of the form US<patent_id>-<sequence_number>, such as "US7315019-0" for the first 
    assignee (Pacific Biosciences) of patent US7315019.
patent_only : Boolean
    If true, only include patent specific information (returns one row). If false, include patent information, 
    CPC categories, and inventor information (returns multiple rows).
connection : sqlalchemy.engine.base.Connection
    Connection using sqlalchemy to the PV database.

Returns
-------
Pandas Dataframe, with columns for (the non-disambiguated version of):
    - URL to the PatentsView website page for the corresponding patent
    - Assignee name
    - Assignee location
    - Co-assignees
    - Patent title
    - Patent classification codes
    - Patent date filed
    - Patent date granted
    - Patent type
    - Patent inventors
    - Inventor locations
    - CPC Subgroups
"""
def assignee_data(mention_id: str, patent_only: bool, connection):
    # Getting patent information
    split = mention_id.split("-")
    patent_id = split[0][2:]
    sequence = split[1]

    # Running query on algorithms_assignee_labeling view
    if patent_only:
        query = f"SELECT patent_id, assignee_sequence, organization, name_first, name_last, assignee_type, title, patent_date,\
            assignee_country, assignee_city, assignee_state FROM algorithms_assignee_labeling.assignee WHERE\
            patent_id='{patent_id}' and assignee_sequence='{sequence}'"
    else:
        query = (f"SELECT * FROM algorithms_assignee_labeling.assignee WHERE patent_id='{patent_id}' and "
                 f"assignee_sequence='{sequence}'")
    result = connection.execute(text(query)).fetchall()

    # Return pandas df
    df = pd.DataFrame(result).drop_duplicates()
    df['assignee_type'] = df['assignee_type'].map(ASSIGNEE_TYPE_DICT)
    return df


"""
Parameters
----------
size : int
    Number of samples

Pulls a random sample of assignee mention_id's from AWS and saves it as an output CSV file
"""
def sample_mentions(size=10000, output_path="data/01 - sample.txt"):
    # Save seed and load in data
    np.random.seed(0)
    disamb = pd.read_csv(
        "g_persistent_assignee.tsv.zip",
        dtype=str,
        sep="\t",
        compression="zip")

    # Clean data
    disamb["mention_id"] = "US" + disamb["patent_id"] + "-" + disamb["assignee_sequence"]
    disamb_20220929 = disamb.set_index("mention_id")["disamb_assignee_id_20230629"]
    disamb_20220929 = disamb_20220929.dropna()
    mention_ids = disamb_20220929.index

    # Save output
    samples = np.random.choice(mention_ids, size=size, replace=False)
    np.savetxt(output_path, samples, fmt="%s")


"""
Parameters
----------
connection : sqlalchemy.engine.base.Connection
    Connection using sqlalchemy to the PV database.
sample_path : str
    Path to CSV file containing a list of mention IDs (`mention_id` field values)
output_path : str
    Path to CSV file for saving populated data. This method both reads and writes for intermitent progress

Output
------
Saves data to `output_path` which is a dataframe with one row for every element in `sample`
Each row has all the attributes in assignee_data()
"""
def populate_sample(connection, sample_path="01 - sample.txt", output_path="02 - sample_with_data.csv"):
    # Load sample data and determine which are previously populated
    sample = np.loadtxt(sample_path, dtype=str)
    prev_df = pd.read_csv(output_path) if os.path.exists(output_path) else pd.DataFrame()
    df_list = [prev_df]
    populated = ["US" + str(row.patent_id) + "-" + str(row.assignee_sequence) for index, row in prev_df.iterrows()]

    for mention_id in sample:
        if mention_id in populated:
            continue

        # Populate mention_id and save data
        temp_df = assignee_data(mention_id, True, connection)
        df_list.append(temp_df)
        populated.append(mention_id)

        # Output messages
        percent = str(round(100 * len(populated) / len(sample), 1)) + "%"
        print(percent, "- created row for", mention_id)

        # Store dataframe intermitently
        if len(populated) % 20 == 0:
            pd.concat(df_list, axis=0, ignore_index=True).to_csv(output_path)

    # Save final dataframe
    pd.concat(df_list, axis=0, ignore_index=True).to_csv(output_path)


"""
Parameters
----------
n (int) : Number of hand labelers
sample_path (str) : path to CSV file containing samples with data
output_path (str) : folder to save n different dataframes

Output
------
Folder with n different files, each an equally sized partition of sample_path
"""
def segment_sample(n=3, sample_path="data/02 - sample_with_data.csv", output_folder="data/03 - segmented samples/"):

    # Load input data and create output folder
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    samples = pd.read_csv(sample_path, index_col=0)
    sample_count = len(samples.index)

    # Values to increment
    increment_size = sample_count // n
    start = 0
    end = increment_size

    # Loop through each hand labeler
    for i in range(n):
        end = end if i in range(sample_count % n) else end - 1 # Handle uneven distribution

        # Save each data partition
        output_path = os.path.join(output_folder, str(i) + " - hand_labeler_file.csv")
        samples.loc[start:end].to_csv(output_path)

        # Increment values for next iteration
        start = end + 1
        end = start + increment_size


"""
Parameters
----------
ws : openpyxl worksheet
    Worksheet for appending rows from 'df' and merging like cells
index_start : int
    First index of the mention_id group
index_end : int
    Last index of the mention_id group
num_columns : int
    Number of columns to apply the merge cell function

Returns
-------
ws : openpyxl worksheet
    Updated worksheet with all rows from `df` to `ws`
"""
def merge_cells(ws, index_start, index_end, num_columns):
    # Loop over columns
    for i in range(1, num_columns + 1):

        # Loop over rows in pre-defined range
        start = index_start
        end = start
        for j in range(index_start, index_end + 1):

            if j == start: # Initialize the previous value at the start of the section
                prev_value = ws.cell(column=i, row=j).value
            elif ws.cell(column=i, row=j).value == prev_value:
                if j == index_end: # Consecutive cells with same value need to merge if end of section
                    ws.merge_cells(start_row=start, end_row=j, start_column=i, end_column=i)
                else: # Otherwise, consecutive cells with same value need to check the next cell before merging
                    end += 1
            elif start != end: # Merge prior sequence once a different value is in place
                ws.merge_cells(start_row=start, end_row=end, start_column=i, end_column=i)
                start = j + 1
                end = j + 1


"""
Parameters
----------
assignee_disamiguation_IDs : list[str]
    List of disambiguated assignee IDs (`assignee_id` field values)
output_path : str
    Path to CSV file for saving data for one mention ID (naming convention is simply the mention ID)
Connection : sqlalchemy.engine.base.Connection
    Connection using sqlalchemy to the PV database.

Processing
----------
1. Pull `df` with SQL connection
2. Group-by mention_id in `df` then sort by inventor_id, and cpc_subgroup_id within the groups
3. Going group by group, add all rows from `df` to `ws`
4. After adding each group, merge all cells that share a value within each column (need to be neighboring cells)
    - Make sure to not merge cells across separate mention_id groups

Returns
-------
openpyxl Workbook with rows for all assignee mentions that correspond to one disambiguated assignee ID in the provided list. 
The columns should be the same as in the `assignee_data` function.
Specifically, rows should be of the form `assignee_data(mention_id)` for each mention_id that corresponds to one of 
the disambiguated assignee ID in the provided list.
Implementing merged cells to increase readability
"""
def disambiguated_assignees_data(assignee_disambiguation_IDs: list[str], connection):
    id_list = '("' + '","'.join(assignee_disambiguation_IDs) + '")'
    query = f"SELECT * FROM algorithms_assignee_labeling.assignee a WHERE a.disambiguated_assignee_id IN {id_list}"
    result = connection.execute(text(query)).fetchall()
    df = pd.DataFrame(result).drop_duplicates()

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Loop through mention_id groups and sort values
    for group, data in df.groupby(['patent_id', 'assignee_sequence']):
        df_temp = data.sort_values(by=["inventor_sequence", "cpc_subgroup_id"], ascending=True, inplace=False)

        # Determine indexing values and add rows to worksheets
        index_start = ws.max_row + 1
        index_end = index_start + len(df_temp.index) - 1
        for row in dataframe_to_rows(df_temp, index=False, header=(index_start == 2)):
            ws.append(row)

        merge_cells(ws, index_start, index_end, len(df.columns))

    return wb

def main():
    engine = establish_connection()
    with engine.connect() as connection:
        # populate_sample('data/sample.csv', 'data/samples_with_data.csv', connection)

        ids = ['717e8394-c25d-4ea6-b444-09d6036a4cde']
        wb = disambiguated_assignees_data(ids, connection)
        output_path = 'data/disamb_assignee_test.xlsx'
        wb.save(output_path)
        print("Successfully saved", output_path)


if __name__ == "__main__":
    main()